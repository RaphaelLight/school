from tkinter import *
from tkinter import ttk, filedialog
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import pandas as pd
import ttkbootstrap as tb
from ttkbootstrap.constants import *
from itemset import Itemset
from associationrule import AssociationRule


main = tb.Window(themename="cosmo")
main.state("zoomed")
main.title("Info Explorer")
main.iconbitmap("Icons\\search.ico")


### FUNCTIONS ###

def fileOpen():
    File = filedialog.askopenfilename(initialdir = "Data\\", title = "Open File", filetype = (("Excel File", "*.xlsx"), ("CSV File", "*.csv")))
    try:
        wb = load_workbook(File)
        ws = wb.active
        heading_row = ws['1']

        global itemset
        itemset = Itemset(File, ws.title)
        global rule
        rule = AssociationRule(File, ws.title)
    except:
        infoScreen.config(text="File couldn't be opened.", font=("courier new", 12), fg="red", justify="left")
        
    itemList.delete(0, END)
    for cell in heading_row:
        itemList.insert(END, cell.value)
    global item_list
    item_list = itemList.get(0, END)

def updateItemList(items):
    itemList.delete(0, END)
    for item in items:
        itemList.insert(END, item)

def checkItemList(e):
    input_= itemSearchBox.get()
    if input_ == "":
        items = item_list
    else:
        items = []
        for item in item_list:
            if input_.lower() in item.lower():
                items.append(item)
    updateItemList(items)

def find_item_set():
    global itemset
    itemset.item_set(itemSetBox.get())

    infoScreen.config(text="Count: " + str(itemset.count()) + "\nSupport: " + str(itemset.support()) + "\nFrequent: " + str(itemset.frequent_item_set(float(itemSetSuppSpin.get()))), font=("courier new", 11), fg="black", anchor="w", )

    resultTable.delete(*resultTable.get_children())

    resultTable.config(columns=("Transactions"), show="headings")
    resultTable.column("Transactions", anchor=CENTER, stretch=YES, minwidth=100)
    resultTable.heading("#0", text="")
    resultTable.heading("Transactions", text="Transactions Containing This Itemset", anchor=CENTER)

    resultTable.tag_configure("oddrow", background="white")
    resultTable.tag_configure("evenrow", background="#EFEFEF")

    rowIndex = 0
    for transaction in itemset.transactionList:
        items = ", ".join(transaction)
        if rowIndex % 2 == 0:
            resultTable.insert(parent="", index="end", values=(items), tags=("evenrow"))
        else:
            resultTable.insert(parent="", index="end", values=(items), tags=("oddrow"))
        rowIndex += 1

    tableStyle = ttk.Style()
    tableStyle.configure("Treeview", background="white", foreground="black", rowheight=20)
    tableStyle.map("Treeview", background=[("selected", "#296DFF")])
    resultTable.config(style=tableStyle)

def plot_support_graph():
    global rule, support_list
    if toggleMode.get() == 1:
        ruleDict = rule.auto_find(float(assocRulSuppSpin.get()), int(minComboSpin.get()))

        for suppCanva in suppGraphFrame.pack_slaves():
            suppCanva.destroy()

        supportFig, supportAx = plt.subplots()
        n = 1
        for key in ruleDict:
            supportAx.bar(n, ruleDict[key], label=key)
            n += 1
        supportFig.set_figwidth(4.9)
        supportFig.set_figheight(3.05)
        supportAx.set_title("Rules by Support")
        supportAx.set_ylabel("Support")
        supportAx.legend()
        suppCanvas = FigureCanvasTkAgg(supportFig, suppGraphFrame)
        suppCanvas.draw()
        suppCanvas.get_tk_widget().pack()
        suppGraphFrame.pack(padx=3, pady=3)
    else:
        supportFig, supportAx = plt.subplots()
        n = 0
        for transaction in rule.transaction_list():
            items = ", ".join(transaction)
            for index in range(n, len(support_list)):
                supportAx.bar(index, support_list[index], label=items)
                n += 1
                break
        supportFig.set_figwidth(4.9)
        supportFig.set_figheight(3.05)
        supportAx.set_title("Rules by Support")
        supportAx.set_ylabel("Support")
        supportAx.legend()
        suppCanvas = FigureCanvasTkAgg(supportFig, suppGraphFrame)
        suppCanvas.draw()
        suppCanvas.get_tk_widget().pack()
        suppGraphFrame.pack(padx=3, pady=3)

def plot_confidence_graph():
    global rule, confidence_list
    confidenceFig, confidenceAx = plt.subplots()
    n = 0
    for transaction in rule.transaction_list():
        items = ", ".join(transaction)
        for index in range(n, len(confidence_list)):
            confidenceAx.bar(index, confidence_list[index], label=items)
            n += 1
            break
    confidenceFig.set_figwidth(4.9)
    confidenceFig.set_figheight(3.05)
    confidenceAx.set_title("Rules by Confidence")
    confidenceAx.set_ylabel("Confidence")
    confidenceAx.legend()
    confCanvas = FigureCanvasTkAgg(confidenceFig, confGraphFrame)
    confCanvas.draw()
    confCanvas.get_tk_widget().pack()
    confGraphFrame.pack(padx=3, pady=3)

def clear_graphs():
    for suppCanva in suppGraphFrame.pack_slaves():
        suppCanva.destroy()
    for confCanva in confGraphFrame.pack_slaves():
        confCanva.destroy()

def find_assoc_rule():
    global rule
    if toggleMode.get() == 1:
        ruleDict = rule.auto_find(float(assocRulSuppSpin.get()), int(minComboSpin.get()))

        infoScreen.config(text="Search Complete", font=("courier new", 11), fg="green", anchor="center")

        resultTable.delete(*resultTable.get_children())

        resultTable.config(columns=("Support", "Itemset"), show="headings") 
        resultTable.column("Support", anchor=CENTER, minwidth=15)
        resultTable.column("Itemset", anchor=CENTER, minwidth=100, stretch=YES)
        resultTable.heading("#0", text="")
        resultTable.heading("Support", text="Support", anchor=CENTER)
        resultTable.heading("Itemset", text="Itemset", anchor=CENTER)

        resultTable.tag_configure("oddrow", background="white")
        resultTable.tag_configure("evenrow", background="#EFEFEF")

        rowIndex = 0
        for key in ruleDict:
            if rowIndex % 2 == 0:
                resultTable.insert(parent="", index="end", values=(ruleDict[key], key), tags=("evenrow"))
            else:
                resultTable.insert(parent="", index="end", values=(ruleDict[key], key), tags=("oddrow"))
            rowIndex += 1

        plot_support_graph()

        tableStyle = ttk.Style()
        tableStyle.configure("Treeview", background="white", foreground="black", rowheight=20)
        tableStyle.map("Treeview", background=[("selected", "#296DFF")])
        resultTable.config(style=tableStyle)

    else:
        itemset_X = antecedentBox.get()
        itemset_Y = consequentBox.get()
        rule.item_sets(itemset_X, itemset_Y)

        infoScreen.config(text="Support: " + str(rule.support()) + "\nConfidence: " + str(rule.confidence()) + "\nLift: " + str(rule.lift()) + "\nStrong Association: " + str(rule.strong_association(float(assocRulSuppSpin.get()), float(assocRulConfSpin.get()))), font=("courier new", 11), fg="black", anchor="w")

        resultTable.delete(*resultTable.get_children())

        resultTable.config(columns=("Support", "Confidence", "Lift", "Antecedent", "Arrow", "Consequent"), show="headings")
        resultTable.column("Support", anchor=CENTER, minwidth=15, width=70)
        resultTable.column("Confidence", anchor=CENTER, minwidth=15, width=80)
        resultTable.column("Lift", anchor=CENTER, minwidth=15, width=50)
        resultTable.column("Antecedent", anchor=E, minwidth=100)
        resultTable.column("Arrow", anchor=CENTER, width=50)
        resultTable.column("Consequent", anchor=W, minwidth=90)
        resultTable.heading("#0", text="")
        resultTable.heading("Support", text="Support", anchor=CENTER)
        resultTable.heading("Confidence", text="Confidence", anchor=CENTER)
        resultTable.heading("Lift", text="Lift", anchor=CENTER)
        resultTable.heading("Antecedent", text="Antecedent", anchor=E)
        resultTable.heading("Arrow", text="→", anchor=CENTER)
        resultTable.heading("Consequent", text="Consequent", anchor=W)

        resultTable.tag_configure("oddrow", background="white")
        resultTable.tag_configure("evenrow", background="#EFEFEF")

        global support_list, confidence_list
        support_list = []
        confidence_list = []

        rowIndex = 0
        for transaction in rule.transaction_list():
            items = consequentBox.get().title()
            consequentItems = items.split(',')
            for item in consequentItems:
                transaction.remove(item)
            antecedent = ", ".join(transaction)
            consequent = ", ".join(consequentItems)
            itemset_X = ','.join(transaction)
            rule.item_sets(itemset_X, itemset_Y)
            if rowIndex % 2 == 0:
                resultTable.insert(parent="", index="end", values=(rule.support(), rule.confidence(), rule.lift(), antecedent, "→", consequent), tags=("evenrow"))
                support_list.append(rule.support())
                confidence_list.append(rule.confidence())
            else:
                resultTable.insert(parent="", index="end", values=(rule.support(), rule.confidence(), rule.lift(), antecedent, "→", consequent), tags=("oddrow"))
                support_list.append(rule.support())
                confidence_list.append(rule.confidence())
            rowIndex += 1

        clear_graphs()
        plot_support_graph()
        plot_confidence_graph()

        tableStyle = ttk.Style()
        tableStyle.configure("Treeview", background="white", foreground="black", rowheight=20)
        tableStyle.map("Treeview", background=[("selected", "#296DFF")])
        resultTable.config(style=tableStyle)

def auto_find():
    if toggleMode.get() == 1:
        antecedentLbl.configure(state="disabled")
        antecedentBox.configure(state="disabled")
        consequentLbl.configure(state="disabled")
        consequentBox.configure(state="disabled")
        assocRulConfLbl.configure(state="disabled")
        assocRulConfSldr.configure(state="disabled")
        assocRulConfSpin.configure(state="disabled")
        minComboFrame.pack(side=RIGHT, padx=6)
        minComboLbl.pack(side=LEFT)
        minComboSpin.pack(side=RIGHT)
    else:
        antecedentLbl.configure(state="normal")
        antecedentBox.configure(state="normal")
        consequentLbl.configure(state="normal")
        consequentBox.configure(state="normal")
        assocRulConfLbl.configure(state="normal")
        assocRulConfSldr.configure(state="normal")
        assocRulConfSpin.configure(state="normal")
        minComboFrame.pack_forget()
        minComboLbl.pack_forget()
        minComboSpin.pack_forget()

########## WIDGET DEFINITIONS ##########

# Menu Bar 
mainMenu = Menu(main)
main.config(menu=mainMenu)
fileMenu = Menu(mainMenu, tearoff=False)
helpMenu = Menu(mainMenu)
mainMenu.add_cascade(label="File", menu=fileMenu)
mainMenu.add_cascade(label="Help", menu=helpMenu)
fileMenu.add_command(label="Open File", command=fileOpen)
fileMenu.add_command(label="Exit")
helpMenu.add_command(label="How To Use")
helpMenu.add_command(label="About")

### FIND SECTION ###
findFrame = LabelFrame(main, text="Find", padx=2, pady=1)
findFrame.grid(row=0, column=0, rowspan=1, padx=3, pady=2, sticky=N)

# Itemset Subsection #
itemSetFrame = LabelFrame(findFrame, text="Itemset", padx=3, pady=3)
findItemSetLbl = Label(itemSetFrame, text="Contains: ")
itemSetBox = ttk.Entry(itemSetFrame, width=29)
itemSetSuppLbl = Label(itemSetFrame, text="Min. Support: ")
IS_sldrVal = DoubleVar()
itemSetSuppSldr = ttk.Scale(itemSetFrame, from_=0, to=1, orient="horizontal", length=133, variable=IS_sldrVal)
itemSetSuppSpin = Spinbox(itemSetFrame, from_=0, to=1, width=5, textvariable=IS_sldrVal)
findItemSetBtn = ttk.Button(itemSetFrame, text="Find Itemset", command=find_item_set)

itemSetFrame.pack(padx=3, pady=3)
findItemSetLbl.grid(row=0, column=0, padx=3, pady=3, sticky=E)
itemSetBox.grid(row=0, column=1, columnspan=2, padx=3, pady=3)
itemSetSuppLbl.grid(row=1, column=0, padx=3, pady=3)
itemSetSuppSldr.grid(row=1, column=1, padx=3, pady=3, sticky=W)
itemSetSuppSpin.grid(row=1, column=2, padx=3, pady=3)
findItemSetBtn.grid(row=2, column=0, columnspan=3, padx=5, pady=5)

# Association Rule Subsection #
assocRuleFrame = LabelFrame(findFrame, text="Association Rule", padx=5, pady=5)
assocRuleFrame.pack(padx=3, pady=6.5)

# Antecedent Subsection
antecedentFrame = LabelFrame(assocRuleFrame, text="Antecedent", padx=5, pady=5)
antecedentLbl = Label(antecedentFrame, text="Contains: ")
antecedentBox = ttk.Entry(antecedentFrame, width=31)

antecedentFrame.pack(padx=3, pady=3)
antecedentLbl.pack(side=LEFT)
antecedentBox.pack(side=RIGHT)

# Consequent Subsection
consequentFrame = LabelFrame(assocRuleFrame, text="Consequent", padx=5, pady=5)
consequentLbl = Label(consequentFrame, text="Contains: ")
consequentBox = ttk.Entry(consequentFrame, width=31)

consequentFrame.pack(padx=3, pady=3)
consequentLbl.pack(side=LEFT)
consequentBox.pack(side=RIGHT)

# Threshold Subsection
thresholdFrame = LabelFrame(assocRuleFrame, text="Threshold", padx=3, pady=3)
assocRulSuppLbl = Label(thresholdFrame, text="Min. Support: ")
AR_SuppSldr = DoubleVar()
assocRulSuppSldr = ttk.Scale(thresholdFrame, from_=0, to=1, orient="horizontal", length=96, variable=AR_SuppSldr)
assocRulSuppSpin = Spinbox(thresholdFrame, from_=0, to=1, width=5, textvariable=AR_SuppSldr)
assocRulConfLbl = Label(thresholdFrame, text="Min. Confidence: ")
AR_ConfSldr = DoubleVar()
assocRulConfSldr = ttk.Scale(thresholdFrame, from_=0, to=1, orient="horizontal", length=96, variable=AR_ConfSldr)
assocRulConfSpin = Spinbox(thresholdFrame, from_=0, to=1, width=5, textvariable=AR_ConfSldr)

thresholdFrame.pack(padx=3, pady=3)
assocRulSuppLbl.grid(row=0, column=0, padx=3, pady=3, sticky=E)
assocRulSuppSldr.grid(row=0, column=1, padx=3, pady=3, sticky=W)
assocRulSuppSpin.grid(row=0, column=2, padx=3, pady=3)
assocRulConfLbl.grid(row=1, column=0, padx=3, pady=3, sticky=E)
assocRulConfSldr.grid(row=1, column=1, padx=3, pady=3, sticky=W)
assocRulConfSpin.grid(row=1, column=2, padx=3, pady=3)

autofindFrame = LabelFrame(thresholdFrame, padx=3, pady=3, relief="flat")
toggleMode = IntVar()
autofindToggle = tb.Checkbutton(autofindFrame, text="Autofind", variable=toggleMode, command=auto_find, bootstyle="primary, round-toggle")

minComboFrame = LabelFrame(autofindFrame, padx=2, pady=2)
minComboLbl = Label(minComboFrame, text="Min. Combo: ")
minComboSpin = Spinbox(minComboFrame, from_=0, to=1000, width=5)

autofindFrame.grid(row=2, column=0, columnspan=3, padx=2, pady=2, sticky=W)
autofindToggle.pack(side=LEFT)

findAssocRuleBtn = ttk.Button(assocRuleFrame, text="Find Association Rule", command=find_assoc_rule)
findAssocRuleBtn.pack(padx=5, pady=5)

### ITEM SECTION ###
itemFrame = LabelFrame(main, text="Items", padx=5, pady=5)
itemSearchBox = Entry(itemFrame, width=41)
itemListFrame = LabelFrame(itemFrame, padx=5, pady=5)
itemScroll_V = Scrollbar(itemListFrame, orient="vertical")
#itemScroll_H = Scrollbar(itemListFrame, orient="horizontal")
itemList = Listbox(itemListFrame, width=42, height=8, yscrollcommand=itemScroll_V.set, exportselection=0)
itemScroll_V.config(command=itemList.yview)
#itemScroll_H.config(command=itemList.xview)
item_list = ()
itemSearchBox.bind("<KeyRelease>", checkItemList)

itemFrame.grid(row=1, column=0, rowspan=1, padx=3, sticky=N)
itemSearchBox.pack(side=TOP, fill="x", pady=5)
itemListFrame.pack(side=BOTTOM)
itemList.pack(padx=3, pady=3, side=LEFT)
itemScroll_V.pack(side=RIGHT, fill='y')
#itemScroll_H.pack(side=BOTTOM, fill='x')


### RESULT SECTION ###
resultFrame = LabelFrame(main, text="Result", width=538, height=465, padx=5, pady=5)
resultScroll_V = Scrollbar(resultFrame, orient="vertical")
resultScroll_H = Scrollbar(resultFrame, orient="horizontal")
resultTable = ttk.Treeview(resultFrame, yscrollcommand=resultScroll_V.set, xscrollcommand=resultScroll_H.set)
resultScroll_V.config(command=resultTable.yview)
resultScroll_H.config(command=resultTable.xview)

resultFrame.grid(row=0, column=1, rowspan=1, padx=3, pady=3, sticky=N)
resultTable.place(width=510, height=420)
resultScroll_V.place(x=510, height=420)
resultScroll_H.place(y=420, width=510)

### INFO SECTION ###
infoFrame = LabelFrame(main, text="Info", padx=5, pady=5, width=536, height=209)
infoScreen = Message(infoFrame, width=520, padx=3, pady=3, bg="white", relief="sunken", text="")

infoFrame.grid(row=1, column=1, padx=3, sticky=N)
infoScreen.place(width=520, height=178)

### SUPPORT GRAPH SECTION ###
graphFrame = LabelFrame(main, relief="flat")
graphFrame.grid(row=0, column=2, rowspan=2, padx=3, pady=3, sticky=N)

suppGraphFrame = LabelFrame(graphFrame, text="Support Analysis", width=500, height=330, padx=3, pady=3)
confGraphFrame = LabelFrame(graphFrame, text="Confidence Analysis", width=500, height=330, padx=3, pady=3)

main.mainloop()