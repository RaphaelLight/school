"""
File: itemset.py
Resources to perform calculations for an itemset.
"""

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

class Itemset(object):

    def __init__(self, dataFile, workSheet):
        wb = load_workbook(dataFile)
        self.ws = wb[workSheet]
        heading_row = self.ws['1']
        self.itemList = []
        for cell in heading_row:
            self.itemList.append(cell.value)
    
    def transaction_set_bin(self):
        key = 1
        rows = self.ws.iter_rows(min_row=key+1, min_col=1)
        transactionSetBin = {}
        for row in rows:
            transaction = []
            for cell in row:
                transaction.append(cell.value)
            transactionSetBin[key] = transaction
            key += 1
        return transactionSetBin
    
    def transaction_set(self):
        transactionSetBin = self.transaction_set_bin()
        transactionSet = []
        for key in transactionSetBin:
            transaction = []
            for n in range(0, len(self.itemList)):
                if transactionSetBin[key][n] == 1:
                    transaction.append(self.itemList[n])
            transactionSet.append(transaction)
        return transactionSet

    def item_set(self, itemSet):
        if not itemSet.__contains__(','):
            self.itemset = []
            self.itemset.append(itemSet)
        elif itemSet.__contains__(', '):
            self.itemset = itemSet.split(', ')
        else:
            self.itemset = itemSet.split(',')

        self.itemset_bin = [0,0,0,0,0]
        for item in self.itemset:
            for product in self.itemList:
                if item.lower() == product.lower():
                    self.itemset_bin[self.itemList.index(product)] = 1
    
    def count(self):
        count = 0
        itemIndices = []
        self.transactionList = []
        for index in range(0, len(self.itemset_bin)):
            if self.itemset_bin[index] == 1:
                itemIndices.append(index)
        for key in self.transaction_set_bin():
            contains = True
            for n in itemIndices:
                if self.itemset_bin[n] != self.transaction_set_bin()[key][n]:
                    contains = False
                    break
            if contains == True:
                count += 1
                self.transactionList.append(self.transaction_set()[key-1])
        return count
    
    def support(self):
        support = float(self.count()/len(self.transaction_set_bin().keys()))
        return support
    
    def frequent_item_set(self, supportThreshold):
        if self.support() >= supportThreshold:
            return True
        else:
            return False

    def __str__(self):
        return str(self.itemset)

        
